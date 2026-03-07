
import openpyxl
wb = openpyxl.load_workbook(r"d:/01_SBS_Projer_GmbH/00_Entwicklung/SBS Projer DEV/Datenanalyse/SBS_Projer_Hauptexcel.xlsm", keep_vba=True, data_only=True)

sheets = ["F_Stoerung", "F_Eigenauftrag", "F_EE_Reinigung", "F_Montage", "F_Pikett", "F_Pauschale"]
target = ["F_Störung", "F_Eigenauftrag", "F_EE_Reinigung", "F_Montage", "F_Pikett", "F_Pauschale"]

print("Available sheets:", wb.sheetnames)

for sheet_name in target:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print()
        print("="*60)
        print(f"SHEET: {sheet_name}")
        print(f"Dimensionen: {ws.dimensions}")
        print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")
        print("Nicht-leere Zellen:")
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    print(f"  Zeile {cell.row}, Spalte {cell.column} ({cell.column_letter}): {repr(cell.value)}")
                    count += 1
        print(f"Total nicht-leere Zellen: {count}")
    else:
        print(f"Sheet not found: {sheet_name}")
