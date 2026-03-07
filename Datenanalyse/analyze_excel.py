import pandas as pd
import openpyxl
from pathlib import Path

# Pfad zur Excel-Datei
excel_file = Path(r"d:\01_SBS_Projer_GmbH\00_Entwicklung\SBS Projer DEV\Datenanalyse\SBS_Projer_Hauptexcel.xlsm")

# Lade das Excel-Workbook
wb = openpyxl.load_workbook(excel_file, keep_vba=True, data_only=True)

print("=" * 80)
print("EXCEL-DATEI ANALYSE: SBS_Projer_Hauptexcel.xlsm")
print("=" * 80)
print()

# Liste alle Tabellenblätter auf
print(f"Anzahl Tabellenblätter: {len(wb.sheetnames)}")
print()
print("Tabellenblätter:")
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {sheet_name}")
print()
print("=" * 80)
print()

# Analysiere jedes Tabellenblatt
for sheet_name in wb.sheetnames:
    print(f"\n{'='*80}")
    print(f"TABELLENBLATT: {sheet_name}")
    print(f"{'='*80}\n")

    try:
        # Lade das Blatt mit pandas
        df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl')

        # Grundlegende Informationen
        print(f"Dimensionen: {df.shape[0]} Zeilen × {df.shape[1]} Spalten")
        print()

        # Spaltennamen
        print("Spalten:")
        for i, col in enumerate(df.columns, 1):
            # Prüfe ob Spalte Daten enthält
            non_null = df[col].notna().sum()
            dtype = df[col].dtype
            print(f"  {i}. {col} ({dtype}) - {non_null} gefüllte Zellen")
        print()

        # Zeige erste paar Zeilen (nur wenn Daten vorhanden)
        if not df.empty and df.shape[0] > 0:
            print("Erste 5 Zeilen (Vorschau):")
            print("-" * 80)
            # Begrenze die Anzahl der angezeigten Spalten für bessere Lesbarkeit
            pd.set_option('display.max_columns', 10)
            pd.set_option('display.width', 200)
            pd.set_option('display.max_colwidth', 30)
            print(df.head().to_string())
            print()

            # Wenn viele Zeilen, zeige auch letzte paar Zeilen
            if df.shape[0] > 10:
                print(f"\nLetzte 3 Zeilen (von {df.shape[0]} gesamt):")
                print("-" * 80)
                print(df.tail(3).to_string())
                print()
        else:
            print("(Keine Daten oder leeres Blatt)")
            print()

        # Statistiken für numerische Spalten
        numeric_cols = df.select_dtypes(include=['number']).columns
        if len(numeric_cols) > 0:
            print("\nStatistiken (numerische Spalten):")
            print("-" * 80)
            print(df[numeric_cols].describe().to_string())
            print()

    except Exception as e:
        print(f"Fehler beim Lesen des Blatts: {e}")
        print()

print("\n" + "=" * 80)
print("ANALYSE ABGESCHLOSSEN")
print("=" * 80)
