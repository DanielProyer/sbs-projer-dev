import pandas as pd
import openpyxl
from pathlib import Path
import json

# Pfad zur Excel-Datei
excel_file = Path(r"d:\01_SBS_Projer_GmbH\00_Entwicklung\SBS Projer DEV\Datenanalyse\SBS_Projer_Hauptexcel.xlsm")

# Lade das Excel-Workbook
wb = openpyxl.load_workbook(excel_file, keep_vba=True, data_only=True)

print("=" * 100)
print("EXCEL-DATEI STRUKTUR-ANALYSE: SBS_Projer_Hauptexcel.xlsm")
print("=" * 100)
print()

# Sammle Struktur-Informationen
structure = {}

for sheet_name in wb.sheetnames:
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl')

        # Sammle Spalteninformationen
        columns_info = []
        for col in df.columns:
            non_null = df[col].notna().sum()
            dtype = str(df[col].dtype)
            columns_info.append({
                'name': str(col),
                'type': dtype,
                'filled': non_null
            })

        structure[sheet_name] = {
            'rows': df.shape[0],
            'cols': df.shape[1],
            'columns': columns_info,
            'has_data': not df.empty and df.shape[0] > 0
        }
    except Exception as e:
        structure[sheet_name] = {
            'error': str(e)
        }

# Ausgabe der Struktur
print(f"Gesamtanzahl Tabellenblätter: {len(wb.sheetnames)}\n")

# Gruppiere Blätter nach Funktion
wochentage = ['Mo', 'Di', 'Mi', 'Do', 'Fr', 'Sa']
auftragstypen = ['Störung', 'Reinigung', 'Eigenauftrag', 'Pikett', 'EE_Reinigung', 'Montage']
formulare = [s for s in wb.sheetnames if s.startswith('F_')]
buchhaltung = ['Journal', 'Hauptbuch', 'Bilanz', 'Erfolgsrechnung', 'Geschäftsfälle', 'Kontenrahmen']
stammdaten = ['Kontakt', 'Betrieb', 'Leitung', 'Anlage', 'Artikel Heineken', 'Listen']

def print_sheet_group(title, sheets):
    print(f"\n{'='*100}")
    print(f"{title}")
    print(f"{'='*100}")
    for sheet in sheets:
        if sheet in structure:
            info = structure[sheet]
            if 'error' in info:
                print(f"\n❌ {sheet}: FEHLER - {info['error']}")
            else:
                rows = info['rows']
                cols = info['cols']
                has_data = "[X]" if info['has_data'] else "[ ]"
                print(f"\n{has_data} {sheet}: {rows} Zeilen × {cols} Spalten")
                if info['columns']:
                    print(f"   Spalten:")
                    for col_info in info['columns']:
                        name = col_info['name']
                        dtype = col_info['type']
                        filled = col_info['filled']
                        if not name.startswith('Unnamed'):
                            print(f"      • {name} ({dtype}) - {filled} Einträge")

# Ausgabe nach Gruppen
print_sheet_group("WOCHENPLANUNG (Routen/Services)", wochentage)
print_sheet_group("AUFTRAGSTYPEN", auftragstypen)
print_sheet_group("FORMULARE", formulare)
print_sheet_group("BUCHHALTUNG", buchhaltung)
print_sheet_group("STAMMDATEN", stammdaten)

# Alle anderen Blätter
andere = [s for s in wb.sheetnames if s not in wochentage + auftragstypen + formulare + buchhaltung + stammdaten]
if andere:
    print_sheet_group("WEITERE BLAETTER", andere)

print("\n" + "=" * 100)
print("ANALYSE ABGESCHLOSSEN")
print("=" * 100)

# Speichere Struktur als JSON
output_json = Path(r"d:\01_SBS_Projer_GmbH\00_Entwicklung\SBS Projer DEV\Datenanalyse\excel_struktur.json")
with open(output_json, 'w', encoding='utf-8') as f:
    json.dump(structure, f, indent=2, ensure_ascii=False)

print(f"\n[OK] Detaillierte Struktur gespeichert in: excel_struktur.json")
